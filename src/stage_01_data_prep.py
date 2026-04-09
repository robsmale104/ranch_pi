"""
stage_01_data_prep.py
Pipeline Stage 1 — Data Preparation
Reads Self_Build_Costs_Data.xlsx, validates, cleans and exports CSVs.
Inputs:  cfg (merged config dict), source Excel file
Outputs: data/clean_inputs.csv, data/gantt_map.csv,
         data/risks.csv, data/assumptions.json
"""
import pandas as pd
import numpy as np
import openpyxl
import json
import os
from pathlib import Path


CBS_GANTT_MAP = {
    'C1.1.1':2,'C1.1.2':3,
    'C1.2.1.1.1':23,'C1.2.1.1.2':23,'C1.2.1.1.3':23,'C1.2.1.1.4':23,'C1.2.1.1.5':23,
    'C1.2.1.2':24,'C1.2.1.3':25,'C1.2.1.4':26,'C1.2.1.5':27,'C1.2.1.6':28,
    'C1.2.1.7':29,'C1.2.1.8':30,'C1.2.1.9':31,
    'C1.2.1.10.1':23,'C1.2.1.10.2':33,'C1.2.1.10.3':33,'C1.2.1.10.4':33,'C1.2.1.10.5':33,
    'C1.2.1.11.1':32,'C1.2.1.11.2':32,'C1.2.1.11.3':32,'C1.2.1.11.4':32,'C1.2.1.11.5':32,
    'C1.2.1.12.1':31,'C1.2.1.12.2':31,'C1.2.1.12.3':31,'C1.2.1.12.4':32,
    'C1.2.1.12.5':31,'C1.2.1.12.6':30,'C1.2.1.12.7':31,
    'C1.2.1.13.1':34,'C1.2.1.13.2':34,'C1.2.1.13.3':34,'C1.2.1.13.4':34,
    'C1.2.1.14.1':48,'C1.2.1.14.2':49,'C1.2.1.14.3':47,
    'C1.2.1.15':50,'C1.2.1.15.1':16,'C1.2.1.16':18,'C1.2.1.17':60,
    'C1.2.1.18.1':61,'C1.2.1.18.2':60,'C1.2.1.18.3':45,'C1.2.1.18.4':45,
    'C1.2.1.19.1':44,'C1.2.1.19.2':45,'C1.2.1.19.3':45,
    'C1.2.1.19.4':45,'C1.2.1.19.5':45,'C1.2.1.19.6':45,
    'C1.2.2.1':36,'C1.2.2.2':36,'C1.2.2.3':37,'C1.2.2.4':38,
    'C1.2.3.1.1':40,'C1.2.3.1.2':40,'C1.2.3.1.3':40,
    'C1.2.3.2.1':40,'C1.2.3.2.2':40,'C1.2.3.2.3':40,
    'C1.2.4.1.1':41,'C1.2.4.1.2':41,'C1.2.4.2':42,'C1.2.4.3':42,
    'C1.2.4.4.1':42,'C1.2.4.4.2':42,'C1.2.4.4.3':42,
    'C1.2.5.1.1':52,'C1.2.5.1.2':52,'C1.2.5.1.3':52,'C1.2.5.2':53,
    'C1.2.5.3.1':54,'C1.2.5.3.2':54,'C1.2.5.3.3':54,'C1.2.5.3.4':54,'C1.2.5.3.5':54,
    'C1.2.5.4.1':55,'C1.2.5.4.2':55,'C1.2.5.4.3':55,
    'C1.2.5.5':56,'C1.2.5.6':56,'C1.2.5.7':57,
    'C1.2.6.1':60,'C1.2.6.2':18,'C1.2.6.3':17,'C1.2.6.4':20,'C1.2.6.5':19,'C1.2.6.6':21,
    'C1.2.7.1':61,'C1.2.7.2':59,'C1.2.7.3':62,
    'C1.3.1':65,'C1.3.2':64,
    'C1.4.1':0,'C1.4.2':0,'C1.4.3':0,
}

GANTT_STREAMS = [
    (2,'PM Appointment & Brief',1,2),
    (3,'Pre-app & Ecology Surveys',1,4),
    (4,'Topographic & Ground Investigation',2,4),
    (5,'Planning Application',2,5),
    (7,'Architectural Design (RIBA 2-3)',3,7),
    (8,'Structural & MEP Engineering',5,9),
    (9,'Contractor Tendering & Appointment',7,11),
    (10,'DNO Grid Connection Application',3,12),
    (11,'Long Lead: Kitchen / Orangery / Steel',6,12),
    (12,'Long Lead: Solar / Wind / Battery',6,11),
    (13,'Long Lead: Greenhouse',7,12),
    (16,'Mains Water Connection',10,13),
    (17,'Borehole & Pump',10,12),
    (18,'Septic Treatment Plant',10,12),
    (19,'Site Drainage & Swales',10,14),
    (20,'Grid Connection Install',12,15),
    (21,'Site Security CCTV & Gates',28,31),
    (23,'Main House — Substructure',11,15),
    (24,'Main House — Superstructure Walls',14,19),
    (25,'Main House — Roof Structure',18,22),
    (26,'Main House — External Envelope',20,25),
    (27,'Main House — First Fix MEP',22,26),
    (28,'Main House — Internal Walls',23,26),
    (29,'Main House — Plaster & Insulation',25,28),
    (30,'Main House — Second Fix MEP',27,30),
    (31,'Main House — Finishes',29,32),
    (32,'Main House — Kitchen & Utility Fitout',30,33),
    (33,'Main House — Basement Fitout',28,33),
    (34,'Main House — Orangery & Outdoor Kitchen',27,31),
    (36,'Annexes — Structure & Roof',14,21),
    (37,'Annexes — Envelope & Fitout',20,28),
    (38,'Annexes — MEP & Finishes',26,31),
    (40,'Garage — Structure Roof & Fitout',12,20),
    (41,'Barns — Foundations & Steel Frame',12,18),
    (42,'Barns — Cladding MEP & Fitout',17,23),
    (44,'Gym — Structure & Shell',14,19),
    (45,'Gym — Fitout',25,32),
    (47,'Energy — Wind Turbine',20,24),
    (48,'Energy — Solar PV',24,27),
    (49,'Energy — Battery Storage',26,28),
    (50,'Energy — Water Collection',25,28),
    (52,'Farm — Fencing & Gates',15,22),
    (53,'Farm — Irrigation & Troughs',18,23),
    (54,'Farm — Animal Housing & Coops',20,26),
    (55,'Farm — Greenhouse & Poly Tunnel',18,24),
    (56,'Farm — Veg Beds & Orchard',24,29),
    (57,'Farm — Equipment Delivery',30,32),
    (59,'Landscape — Wild Pond',20,24),
    (60,'Landscape — Driveway & Paving',22,27),
    (61,'Landscape — Garden',27,33),
    (62,'Landscape — Furniture & Lighting',31,33),
    (64,'Testing — MEP & Utilities Commissioning',31,34),
    (65,'Testing — Building Control',32,35),
    (66,'Testing — Snagging & Handover',34,36),
]


def run(cfg, logger):
    """Main entry point called by pipeline.py"""
    data_dir = Path(cfg["data_dir"])
    data_dir.mkdir(exist_ok=True)

    logger.info("Stage 01 — Data Preparation starting")

    # ── Save assumptions.json ──────────────────────────────────────
    assumptions = {
        "project":           cfg["project"],
        "base_year":         cfg["base_year"],
        "currency":          cfg["currency"],
        "main_house_gia_m2": cfg["main_house_gia_m2"],
        "n_simulations":     cfg["n_simulations"],
        "random_seed":       cfg["random_seed"],
        "inflation_rate":    cfg["inflation_rate"],
        "inflation_base_year": cfg["inflation_base_year"],
        "cpi_factors":       cfg["cpi_factors"],
        "driver_order":      cfg["driver_order"],
        "corr_matrix":       cfg["corr_matrix"],
        "programme_months":  cfg["programme_months"],
        "inservice_start_yr": cfg["inservice_start_yr"],
        "inservice_end_yr":   cfg["inservice_end_yr"],
        "total_model_months": cfg["total_model_months"],
        "inservice_cbs":      cfg["inservice_cbs"],
        "daily_delay_cost_gbp": cfg["daily_delay_cost_gbp"],
        "days_per_month":       cfg["days_per_month"],
        "scenario_name":        cfg.get("scenario_name","Base"),
    }
    with open(data_dir / "assumptions.json", "w") as f:
        json.dump(assumptions, f, indent=2)
    logger.info("Saved data/assumptions.json")

    # ── Load workbook ──────────────────────────────────────────────
    logger.info(f"Loading {cfg['source_file']}")
    wb = openpyxl.load_workbook(cfg["source_file"], read_only=True)

    # ── Cost Lines ────────────────────────────────────────────────
    ws_cl = wb["Cost_Lines"]
    rows  = list(ws_cl.iter_rows(values_only=True))
    df_raw = pd.DataFrame(rows[1:], columns=rows[0])

    col_map = {
        "CBS_Code":"cbs_code","CBS_Description":"description",
        "CBS_Level":"cbs_level","WBS_Phase":"phase","Row_Type":"row_type",
        "Unit":"unit","Qty":"qty","Size_Assumption":"size_note",
        "Unit_Cost_Low (£)":"unit_cost_low","Unit_Cost_ML (£)":"unit_cost_ml",
        "Unit_Cost_High (£)":"unit_cost_high","Total_Low (£)":"total_low",
        "Total_ML (£)":"total_ml","Total_High (£)":"total_high",
        "Duration_Low (days)":"dur_low_days","Duration_ML (days)":"dur_ml_days",
        "Duration_High (days)":"dur_high_days","Notes / Assumptions":"notes",
        "Sensitivity_Drivers":"sens_drivers","Sensitivity_Rationale":"sens_rationale",
    }
    df_raw = df_raw.rename(columns=col_map)
    df_inputs  = df_raw[df_raw["row_type"] == "INPUT"].copy()
    df_rollups = df_raw[df_raw["row_type"] == "ROLLUP"].copy()
    logger.info(f"Cost_Lines: {len(df_inputs)} INPUT, {len(df_rollups)} ROLLUP")

    # ── Expand in-service rows ─────────────────────────────────────
    is_cbs = cfg["inservice_cbs"]
    inservice_base = df_inputs[df_inputs["cbs_code"].isin(is_cbs)].copy()
    build_inputs   = df_inputs[~df_inputs["cbs_code"].isin(is_cbs)].copy()

    inservice_rows = []
    for yr in range(cfg["inservice_start_yr"], cfg["inservice_end_yr"] + 1):
        yr_rows = inservice_base.copy()
        yr_rows["cbs_code"] = yr_rows["cbs_code"] + f"_{yr}"
        yr_rows["phase"]    = f"Phase 5 ({yr})"
        yr_rows["inservice_year"] = yr
        inservice_rows.append(yr_rows)

    df_inservice = pd.concat(inservice_rows, ignore_index=True)
    build_inputs["inservice_year"] = None
    df_inputs = pd.concat([build_inputs, df_inservice], ignore_index=True)

    logger.info(f"In-service expanded: {len(df_inservice)} rows "
                f"({len(is_cbs)} lines x "
                f"{cfg['inservice_end_yr']-cfg['inservice_start_yr']+1} years)")

    # ── Gantt index mapping ────────────────────────────────────────
    def get_gantt_index(cbs_code):
        base = cbs_code.split("_")[0]
        return CBS_GANTT_MAP.get(base, None)

    df_inputs["gantt_index"] = df_inputs["cbs_code"].apply(get_gantt_index)
    unmapped = df_inputs[df_inputs["gantt_index"].isna()]
    if len(unmapped):
        logger.warning(f"{len(unmapped)} rows not mapped to Gantt index")
    else:
        logger.info(f"All {len(df_inputs)} rows mapped to Gantt index")

    # ── Validation ────────────────────────────────────────────────
    build_rows = df_inputs[df_inputs["inservice_year"].isna()]
    cost_rows  = build_rows.dropna(subset=["total_low","total_ml","total_high"])
    bad = cost_rows[
        (cost_rows["total_low"] > cost_rows["total_ml"]) |
        (cost_rows["total_ml"] > cost_rows["total_high"])
    ]
    if len(bad):
        raise RuntimeError(f"{len(bad)} rows violate Low<=ML<=High")

    dupes = build_rows[build_rows["cbs_code"].duplicated()]
    if len(dupes):
        raise RuntimeError(f"{len(dupes)} duplicate CBS codes")

    logger.info("Validation passed: ordering, duplicates OK")

    # ── Save clean inputs ──────────────────────────────────────────
    df_inputs.to_csv(data_dir / "clean_inputs.csv", index=False)
    logger.info(f"Saved data/clean_inputs.csv ({len(df_inputs)} rows)")

    # ── Gantt map ──────────────────────────────────────────────────
    df_gantt = pd.DataFrame(GANTT_STREAMS,
                            columns=["gantt_index","stream_label",
                                     "start_month","end_month"])
    df_gantt["duration_months"] = df_gantt["end_month"] - df_gantt["start_month"] + 1

    is_gantt = []
    for yr in range(cfg["inservice_start_yr"], cfg["inservice_end_yr"] + 1):
        offset  = (yr - cfg["inservice_start_yr"]) * 12
        is_gantt.append({
            "gantt_index":    0,
            "stream_label":   f"In-Service ({yr})",
            "start_month":    37 + offset,
            "end_month":      48 + offset,
            "duration_months": 12,
            "inservice_year": yr,
        })

    df_gantt["inservice_year"] = None
    df_gantt = pd.concat([df_gantt, pd.DataFrame(is_gantt)], ignore_index=True)
    df_gantt.to_csv(data_dir / "gantt_map.csv", index=False)
    logger.info(f"Saved data/gantt_map.csv ({len(df_gantt)} rows)")

    # ── Risks ──────────────────────────────────────────────────────
    ws_r   = wb["Risks"]
    r_rows = list(ws_r.iter_rows(values_only=True))
    df_risks = pd.DataFrame(r_rows[1:], columns=r_rows[0])
    df_risks = df_risks.dropna(subset=["Risk ID"])
    df_risks.columns = ["risk_id","description","likelihood","_raw",
                        "dur_low","dur_ml","dur_high",
                        "cost_low","cost_ml","cost_high","mitigation"]
    df_risks = df_risks.drop(columns=["_raw"])
    df_risks.to_csv(data_dir / "risks.csv", index=False)
    logger.info(f"Saved data/risks.csv ({len(df_risks)} risks)")

    logger.info("Stage 01 complete")
    return {
        "n_inputs":      len(df_inputs),
        "n_build":       len(build_inputs),
        "n_inservice":   len(df_inservice),
        "build_ml_total": float(build_inputs["total_ml"].sum()),
    }
